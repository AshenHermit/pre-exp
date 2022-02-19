from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles.alignment import Alignment
from openpyxl.styles.fonts import Font
from openpyxl.utils import get_column_letter

x_offset = 5
y_offset = 4

class Builder:
    def __init__(self, generation_type="graph"):
        self.generation_type = generation_type
        self.days_count_in_month = 31

    def translate_to_new_input_table(self, input_path, output_path):
        print("reading input workbook...")
        input_workbook = load_workbook(filename=input_path)
        input_sheet = input_workbook.active
        print("done.")
        
        print("reading output workbook...")
        output_workbook = Workbook()
        output_sheet = output_workbook.active
        print("done.")

        # processing
        y = 8
        name = None
        while y==8 or name:
            name = input_sheet["E{}".format(y)].value
            if type(name) is str:
                name = name.lstrip()
                output_sheet.cell(column=1, row=y-8+1).value = name
                for x in range(31):
                    output_sheet.cell(column=x+2, row=y-8+1).value = input_sheet.cell(column=42 + (x*4), row=y).value
                print(name)
            else:
                name = None
            y+=1

        

        # output_sheet.cell(row=10, column=6).value

        output_workbook.save("./output.xlsx")

    def build(self, input_path, output_path):
        print("reading input workbook...")
        input_workbook = load_workbook(filename=input_path)
        input_sheet = input_workbook.active
        print("done.")
        
        print("reading output workbook...")
        #output_workbook = load_workbook(filename=output_path)
        output_workbook = Workbook()
        output_sheet = output_workbook.active
        print("done.")


        # gather data
        staff = []
        y = 6
        name = None
        while y==6 or name:
            name = input_sheet["B{}".format(y)].value
            if type(name) is str:
                name = name.lstrip()

                job = input_sheet["C{}".format(y)].value
                try: job = job.lstrip()
                except: pass

                data = {'name': name, 'job': job, 'days': []}
                for x in range(self.days_count_in_month):
                    day = input_sheet.cell(column=4 + x, row=y).value
                    try: day = day.lstrip().lower()
                    except: pass
                    data['days'].append(day)
                staff.append(data)
            else:
                name = None
            y+=1


        # process data
        staff = sorted(staff, key=lambda x: x['name'])

        # generate output sheet
        ## header
        last_cell_x = 1+self.days_count_in_month+(2+1+1+1)

        output_sheet.merge_cells('A1:A2')
        output_sheet.merge_cells('B1:C2')
        output_sheet.merge_cells('D1:D2')
        output_sheet.merge_cells(start_row=1, start_column=5, end_row=1, end_column=last_cell_x)

        # staff
        for i, emp in enumerate(staff):
            # day / night

            k = (2 if self.generation_type=="table" else 1)

            for dn in range(k):
                y = i*2*k+dn*2
                
                for x in range(1, 4):
                    output_sheet.merge_cells(start_row=y_offset + y, start_column=x, end_row=y_offset + y+1, end_column=x)
            
                # name
                output_sheet.cell(column=1, row=y_offset + y).alignment = Alignment(horizontal='left', vertical='center')
                output_sheet.cell(column=1, row=y_offset + y).value = emp['name']

                # number
                output_sheet.cell(column=2, row=y_offset + y).alignment = Alignment(horizontal='center', vertical='center')
                output_sheet.cell(column=2, row=y_offset + y).value = i+1

                # job
                output_sheet.cell(column=4, row=y_offset + y).alignment = Alignment(horizontal='center', vertical='center')
                output_sheet.cell(column=4, row=y_offset + y).value = emp['job']
                output_sheet.cell(column=4, row=y_offset + y+1).alignment = Alignment(horizontal='center', vertical='center')
                output_sheet.cell(column=4, row=y_offset + y+1).value = 1


        for x in range(x_offset, last_cell_x):
            output_sheet.column_dimensions[get_column_letter(x)].width = 6.4

        


        if self.generation_type == "table":
            self.build_table(staff, output_sheet, last_cell_x)

        elif self.generation_type == "graph":
            self.build_graph(staff, output_sheet, last_cell_x)

        output_workbook.save(output_path)

    def build_graph(self, staff, output_sheet, last_cell_x):
        # main data
        last_cell_x-=1
        for x in range(1, last_cell_x+1): # + Уч. номер(2) + Должность(1) + Итого_1(1) + Всего(1)
            output_sheet.cell(column=x, row=3).font = Font(size=6)

            output_sheet.cell(column=x, row=3).alignment = Alignment(horizontal='center')
            output_sheet.cell(column=x, row=3).value = x

            if x == 1:
                output_sheet.cell(column=x, row=1).alignment = Alignment(horizontal='center')
                output_sheet.cell(column=x, row=1).value = "Фамилия имя отчетсво\n(строго в алфавитном порядке)"

            elif x==2:  
                output_sheet.cell(column=x, row=1).alignment = Alignment(horizontal='center')
                output_sheet.cell(column=x, row=1).value = "Уч. номер"

            elif x==4:
                output_sheet.cell(column=x, row=1).alignment = Alignment(horizontal='center')
                output_sheet.cell(column=x, row=1).value = "Должность (профессия) / Ставка"

            elif x==5:
                output_sheet.cell(column=x, row=1).alignment = Alignment(horizontal='center')
                output_sheet.cell(column=x, row=1).value = "Числа месяца"

            if x>4 and x<last_cell_x:
                output_sheet.cell(column=x, row=2).alignment = Alignment(horizontal='center')
                output_sheet.cell(column=x, row=2).value = x-4

            elif x==last_cell_x:
                output_sheet.cell(column=x, row=2).alignment = Alignment(horizontal='center')
                output_sheet.cell(column=x, row=2).value = "Основная ставка"

        #staff
        for i, emp in enumerate(staff):
            # day / night
            for dn in range(1):
                y = i*2+dn*2

                # days
                for x, day in reversed(list(enumerate(emp['days']))):
                    if dn == 0:
                        dx = x_offset + x
                        dy = y_offset + y
                        
                        if day == "" or day == "*" or day == None:
                            pass
                            # self.set_table_value(output_sheet, dx, dy+1, "В")

                        elif day == "с":
                            self.set_table_value(output_sheet, dx, dy, "7:59")
                            self.set_table_value(output_sheet, dx, dy+1, "23:59")
                            self.set_table_value(output_sheet, dx+1, dy, "0:01")
                            self.set_table_value(output_sheet, dx+1, dy+1, "8:01")

                        elif day=="1,8":
                            self.set_table_value(output_sheet, dx+1, dy, "16:0")
                            self.set_table_value(output_sheet, dx+1, dy+1, "17:48")

                        elif day=="0/8":
                            self.set_table_value(output_sheet, dx+1, dy, "0:01")
                            self.set_table_value(output_sheet, dx+1, dy+1, "8:01")

                        elif str(day).isdigit():
                            if emp['job'][:5] == 'буфет' or emp['job'][:5] == 'уборщ':
                                self.set_table_value(output_sheet, dx, dy, "8:00")
                                self.set_table_value(output_sheet, dx, dy+1, "20:00")

                            else:
                                self.set_table_value(output_sheet, dx, dy, "7:59")
                                self.set_table_value(output_sheet, dx, dy+1, "23:59")
                                self.set_table_value(output_sheet, dx+1, dy, "0:01")
                                self.set_table_value(output_sheet, dx+1, dy+1, "8:01")

                        elif day=="д":
                            self.set_table_value(output_sheet, dx, dy, "8:00")
                            self.set_table_value(output_sheet, dx, dy+1, "15:42")

                        else:
                            pass
                            # self.set_table_value(output_sheet, dx, dy+1, day)




                    

    def build_table(self, staff, output_sheet, last_cell_x):
        # main data
        for x in range(1, last_cell_x+1): # + Уч. номер(2) + Должность(1) + Итого_1(1) + Всего(1)
            output_sheet.cell(column=x, row=3).font = Font(size=6)

            output_sheet.cell(column=x, row=3).alignment = Alignment(horizontal='center')
            output_sheet.cell(column=x, row=3).value = x

            if x == 1:
                output_sheet.cell(column=x, row=1).alignment = Alignment(horizontal='center')
                output_sheet.cell(column=x, row=1).value = "Фамилия имя отчетсво\n(строго в алфавитном порядке)"

            elif x==2:  
                output_sheet.cell(column=x, row=1).alignment = Alignment(horizontal='center')
                output_sheet.cell(column=x, row=1).value = "Уч. номер"

            elif x==4:
                output_sheet.cell(column=x, row=1).alignment = Alignment(horizontal='center')
                output_sheet.cell(column=x, row=1).value = "Должность (профессия) / Ставка"

            elif x==5:
                output_sheet.cell(column=x, row=1).alignment = Alignment(horizontal='center')
                output_sheet.cell(column=x, row=1).value = "Числа месяца"

            if x>4 and x<20:
                output_sheet.cell(column=x, row=2).alignment = Alignment(horizontal='center')
                output_sheet.cell(column=x, row=2).value = x-4

            elif x==20:
                output_sheet.cell(column=x, row=2).alignment = Alignment(horizontal='center')
                output_sheet.cell(column=x, row=2).value = "Итого"

            elif x>20 and x<last_cell_x:
                output_sheet.cell(column=x, row=2).alignment = Alignment(horizontal='center')
                output_sheet.cell(column=x, row=2).value = x-5

            elif x==last_cell_x:
                output_sheet.cell(column=x, row=2).alignment = Alignment(horizontal='center')
                output_sheet.cell(column=x, row=2).value = "Всего"
        
        #staff
        for i, emp in enumerate(staff):
            # day / night
            for dn in range(2):
                y = i*4+dn*2

                # formulas
                output_sheet.cell(column=20, row=y_offset + y).value = "=SUM(E{0}:S{0})".format(y_offset + y)
                if self.days_count_in_month == 30:
                    output_sheet.cell(column=last_cell_x, row=y_offset + y).value = "=SUM(U{0}:AI{0})+T{0}".format(y_offset + y)
                if self.days_count_in_month == 31:
                    output_sheet.cell(column=last_cell_x, row=y_offset + y).value = "=SUM(U{0}:AJ{0})+T{0}".format(y_offset + y)

                # days
                for x, day in reversed(list(enumerate(emp['days']))):
                    if dn == 0:
                        dx = x_offset + x
                        dy = y_offset + y
                        
                        if day == "" or day == "*" or day == None:
                            self.set_table_value(output_sheet, dx, dy+1, "В")

                        elif day == "с":
                            self.set_table_value(output_sheet, dx, dy, 15)
                            self.set_table_value(output_sheet, dx, dy+1, "Ф")
                            self.set_table_value(output_sheet, dx+1, dy, 7.5)
                            self.set_table_value(output_sheet, dx+1, dy+1, "Ф")

                            self.set_table_value(output_sheet, dx, dy+2, 2)
                            self.set_table_value(output_sheet, dx, dy+3, "Н")
                            self.set_table_value(output_sheet, dx+1, dy+2, 6)
                            self.set_table_value(output_sheet, dx+1, dy+3, "Н")

                        elif day=="1,8":
                            self.set_table_value(output_sheet, dx, dy, 1.8)
                            self.set_table_value(output_sheet, dx, dy+1, "СТ")

                        elif day=="0/8":
                            self.set_table_value(output_sheet, dx, dy, 7.5)
                            self.set_table_value(output_sheet, dx, dy+1, "Ф")
                            self.set_table_value(output_sheet, dx, dy+2, 6)
                            self.set_table_value(output_sheet, dx, dy+3, "Н")

                        elif str(day).isdigit():
                            if emp['job'][:5] == 'буфет' or emp['job'][:5] == 'уборщ':
                                self.set_table_value(output_sheet, dx, dy, 11.5)
                                self.set_table_value(output_sheet, dx, dy+1, "Ф")

                            else:
                                self.set_table_value(output_sheet, dx, dy, 12)
                                self.set_table_value(output_sheet, dx, dy+1, "Ф")
                                self.set_table_value(output_sheet, dx+1, dy, 12)
                                self.set_table_value(output_sheet, dx+1, dy+1, "Ф")

                                self.set_table_value(output_sheet, dx, dy+2, 7.2)
                                self.set_table_value(output_sheet, dx, dy+3, "Н")
                                self.set_table_value(output_sheet, dx+1, dy+2, 7.2)
                                self.set_table_value(output_sheet, dx+1, dy+3, "Н")

                        elif day=="д":
                            self.set_table_value(output_sheet, dx, dy, 7.2)
                            self.set_table_value(output_sheet, dx, dy+1, "Ф")

                        else:
                            self.set_table_value(output_sheet, dx, dy+1, day)
    

    def set_table_value(self, sheet, x, y, value):
        if self.generation_type =='table' and x >= 20:
            x+=1
            limit = 1+self.days_count_in_month+(2+1+1+1)
        else:
            limit = self.days_count_in_month+(2+1+1+1)

        if x < limit:
            sheet.cell(column=x, row=y).alignment = Alignment(horizontal='center', vertical='center')
            sheet.cell(column=x, row=y).value = value
