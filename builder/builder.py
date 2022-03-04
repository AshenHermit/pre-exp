from pathlib import Path
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl import load_workbook
from openpyxl.styles.alignment import Alignment
from openpyxl.styles.fonts import Font
from openpyxl.utils import get_column_letter
import typing

from builder.sheet_builder_base import ExcelSheetBuilder

from .data import RulesLibrary, StaffData
from .data_readers import RulesReader, StaffDataReader

class ExcelWorkbookWriter():
    def __init__(self, builder:ExcelSheetBuilder, data_xlsx:Path, output_xlsx:Path, data_sheet_name:str) -> None:
        self.builder:ExcelSheetBuilder = builder
        self.data_xlsx:Path = data_xlsx
        self.output_xlsx:Path = output_xlsx
        self.data_sheet_name:str = data_sheet_name
        
    def run(self):
        self.read_data_file()
        self.generate_to_file(self.builder, self.output_xlsx)

    def read_data_file(self):
        print(f"читаю файл \"{self.data_xlsx.name}\", лист \"{self.data_sheet_name}\"...")
        self.data_wb = load_workbook(str(self.data_xlsx))
        self.data_sheet = self.data_wb[self.data_sheet_name]
        
    def generate_to_file(self, builder:ExcelSheetBuilder, output_file:Path):
        print(f"подготавливаю \"{output_file.name}\"...")
        output_wb = Workbook()
        gen_sheet = output_wb.active

        print(f"генерирую...")
        builder.workbook = output_wb
        builder.data_sheet = self.data_sheet
        builder.sheet = gen_sheet
        builder.build()

        print(f"сохраняю \"{output_file.name}\"...")
        output_wb.save(str(output_file))
        
