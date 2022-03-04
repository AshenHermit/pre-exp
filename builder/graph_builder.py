import datetime
from pathlib import Path
import traceback
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl import load_workbook
from openpyxl.styles.alignment import Alignment
from openpyxl.styles.fonts import Font
from openpyxl.utils import get_column_letter
import typing

from . import utils

from builder.builder import ExcelSheetBuilder

from .data import RulesLibrary, StaffData
from .data_readers import RulesReader, StaffDataReader

class GraphBuilder(ExcelSheetBuilder):
    def render_base(self):
        self.render_people_list(name_repeat=1)
        self.render_titles()

    def render_titles(self):
        # main data
        self.last_cell_x-=1
        for x in range(1, self.last_cell_x+2): # + Уч. номер(2) + Должность(1) + Итого_1(1) + Всего(1)
            self.sheet.cell(column=x, row=3).font = Font(size=6)

            self.sheet.cell(column=x, row=3).alignment = Alignment(horizontal='center')
            self.sheet.cell(column=x, row=3).value = x

            if x == 1:
                self.sheet.cell(column=x, row=1).alignment = Alignment(horizontal='center')
                self.sheet.cell(column=x, row=1).value = "Фамилия имя отчетсво\n(строго в алфавитном порядке)"

            elif x==2:  
                self.sheet.cell(column=x, row=1).alignment = Alignment(horizontal='center')
                self.sheet.cell(column=x, row=1).value = "Уч. номер"

            elif x==4:
                self.sheet.cell(column=x, row=1).alignment = Alignment(horizontal='center')
                self.sheet.cell(column=x, row=1).value = "Должность (профессия) / Ставка"

            elif x==5:
                self.sheet.cell(column=x, row=1).alignment = Alignment(horizontal='center')
                self.sheet.cell(column=x, row=1).value = "Числа месяца"

            if x>4 and x<self.last_cell_x:
                self.sheet.cell(column=x, row=2).alignment = Alignment(horizontal='center')
                self.sheet.cell(column=x, row=2).value = x-4

            elif x==self.last_cell_x:
                self.sheet.cell(column=x, row=2).alignment = Alignment(horizontal='center')
                self.sheet.cell(column=x, row=2).value = "Основная ставка"
                
            elif x==self.last_cell_x+1:
                self.sheet.cell(column=x, row=2).alignment = Alignment(horizontal='center')
                self.sheet.cell(column=x, row=2).value = "Ночь"
    
    def fill_data(self):
        #staff
        for i, emp in enumerate(self.staff.people):
            # day / night
            for dn in range(1):
                y = i*2+dn*2

                # days
                total_day = 0
                total_night = 0
                dx = 0
                dy = 0
                for x, day in reversed(list(enumerate(emp.days))):

                    if dn != 0: continue

                    dx = self.data_start_x + x
                    dy = self.data_start_y + y
                    
                    if day == None:
                        day = ""
                    
                    rule = self.rules.get_fit_rule(day, emp.job)
                    if rule is not None:
                        graph_table = rule.tables[1]
                        for i in range(graph_table.size[0]):
                            for j in range(graph_table.size[1]):
                                self.set_table_value(dx+i, dy+j, graph_table.get(i, j))
                        
                        exp_table = rule.tables[0]
                        for i in range(exp_table.size[0]):
                            for j in range(exp_table.size[1]):
                                try:
                                    value = float(exp_table.get(i,j, 0))
                                    if j==0: total_day += value
                                    if j==2: total_night += value
                                except:
                                    pass
                
                self.set_table_value(
                    self.data_start_x+self.days_count_in_month, self.data_start_y+y, 
                    total_day, False)
                self.set_table_value(
                    self.data_start_x+self.days_count_in_month+1, self.data_start_y+y, 
                    total_night, False)

    def set_table_value(self, x, y, value, use_limit=True):
        value = self.format_cell_value(value)
        limit = self.days_count_in_month + 5

        if x < limit or not use_limit:
            self.sheet.cell(column=x, row=y).alignment = Alignment(horizontal='center', vertical='center')
            self.sheet.cell(column=x, row=y).value = value