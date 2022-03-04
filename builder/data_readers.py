from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl import load_workbook
from openpyxl.styles.alignment import Alignment
from openpyxl.styles.fonts import Font
from openpyxl.utils import get_column_letter
import typing
import shlex
import datetime

from .data import PersonData, Rule, RulesLibrary, SimpleTable, StaffData

# классы для чтения листа (sheet) с сырыми данными, например лист "ВСЕ"

class DataReader():
    def __init__(self, sheet:Worksheet) -> None:
        self.sheet:Worksheet = sheet

class StaffDataReader(DataReader):
    def __init__(self, sheet:Worksheet):
        super().__init__(sheet)
        self.days_count_in_month = 31

    def read_staff_data(self):
        staff = StaffData()

        y = 6
        empty_space_i = 0
        while y==6 or empty_space_i < 10:
            name = self.sheet["B{}".format(y)].value
            if type(name) is str:
                name = name.lstrip()
                empty_space_i = 0

                job = self.sheet["C{}".format(y)].value
                try: job = job.lstrip()
                except: pass

                person = PersonData(name, job, [])
                for x in range(self.days_count_in_month):
                    day = self.sheet.cell(column=4 + x, row=y).value
                    try: day = day.lstrip().lower()
                    except: pass
                    person.days.append(day)
                staff.add_person(person)
            else:
                empty_space_i += 1
            y+=1

        staff.sort()
        
        return staff

class RulesReader(DataReader):
    def __init__(self, sheet:Worksheet):
        super().__init__(sheet)
        self.days_count_in_month = 31

        self.table_names_y = 5
        self.rules_start_x = 4 + self.days_count_in_month

        self.y = self.table_names_y+1
        self.x = self.rules_start_x
        self.last_rule_name = ""
        self.last_table_name = ""

    def read_rule_library(self):
        rlib = RulesLibrary()
        
        self.y = self.table_names_y+1
        self.x = self.rules_start_x
        self.last_rule_name = ""
        self.last_table_name = str(self.sheet.cell(column=self.x, row=self.table_names_y).value)
        iter = 0
        while True:
            next_rule_y = self.find_next_rule_y()
            if iter==0: next_rule_y = self.y
            if next_rule_y==-1: break
            self.y = next_rule_y
            rule_name = str(self.read_val())
            self.last_rule_name = rule_name
            
            rule = Rule()
            rule.day_values = shlex.split(rule_name.replace(",", " "))
            self.read_tables(rule)
            rlib.rules.append(rule)
            iter+=1
        
        return rlib

    def find_next_rule_y(self):
        prev_value = "- --prev _val ue-- -"
        for i in range(8):
            i = i+1
            value = self.read_val(y_offset=i)
            if prev_value is None and value is not None:
                value = str(value)
                return self.y+i
            prev_value = value
        return -1

    def find_next_table_x(self):
        for i in range(8):
            value = self.read_table_name(x_offset=i)
            if value is not None:
                value = str(value)
                if value != self.last_table_name:
                    return self.x+i
        return -1
    
    def read_tables(self, rule:Rule):
        next_rule_y = self.find_next_rule_y()
        max_y_off = next_rule_y-self.y if next_rule_y != -1 else 8

        self.last_table_name = self.read_table_name()
        self.x = self.find_next_table_x()
        self.last_table_name = self.read_table_name()
        
        while True:
            next_table_x = self.find_next_table_x()
            max_x_off = next_table_x-self.x if next_table_x != -1 else 8
            table = SimpleTable.from_sheet_region(self.sheet, (self.x, self.y), (self.x+max_x_off, self.y+max_y_off))
            rule.tables.append(table)
            self.x += max_x_off
            if next_table_x != -1:
                self.last_table_name = self.read_table_name()
            else:
                break
        
        self.x = self.rules_start_x

    def read_table_name(self, x_offset=0):
        return self.sheet.cell(column = self.x+x_offset, row=self.table_names_y).value
    
    def read_val(self, x_offset=0, y_offset=0):
        return self.sheet.cell(column = self.x+x_offset, row = self.y+y_offset).value

class DateInfo():
    def __init__(self) -> None:
        self.str_date = ""
class DateReader(DataReader):
    def __init__(self, sheet: Worksheet):
        super().__init__(sheet)

    def read_date_info(self):
        info = DateInfo()
        info.str_date = str(self.sheet.cell(column=1, row=1).value)
        return info