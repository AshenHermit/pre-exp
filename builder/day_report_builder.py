import datetime
from functools import cache
from pathlib import Path
import traceback
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl import load_workbook
from openpyxl.styles.alignment import Alignment
from openpyxl.styles.fonts import Font
from openpyxl.utils import get_column_letter
import typing

from . import utils

from builder.builder import ExcelSheetBuilder

from .data import PersonData, RulesLibrary, StaffData
from .data_readers import RulesReader, StaffDataReader

class DayReportBuilder(ExcelSheetBuilder):
    def __init__(self, data_sheet: Worksheet = None, sheet: Worksheet = None, selected_day:int=0):
        super().__init__(data_sheet, sheet)
        self.selected_day = selected_day

    @property
    def group_people_map(self):
        group_people_map:dict[str, list] = {}
        for person in self.staff.people:
            day = person.days[self.selected_day-1] or ""
            rule = self.rules.get_fit_rule(day, person.job)
            if rule is not None:
                for group_name in rule.groups_names:
                    if group_name not in group_people_map:
                        group_people_map[group_name] = []
                    group_people_map[group_name].append(person)
        return group_people_map

    def build(self):
        self.setup_data()

        self.workbook.remove(self.workbook.active)
        self.build_handy_sheet()

    def build_handy_sheet(self):
        self.sheet = self.workbook.create_sheet("список")
        self.make_handy_sheet()
        self.sheet = self.workbook.create_sheet("отчет")
        self.make_report_sheet()

    def make_handy_sheet(self):
        self.data_start_x = 2
        self.data_start_y = 3
        group_people_map = self.group_people_map

        end_col = 1+len(group_people_map)
        for i in range(2):
            self.sheet.merge_cells(
                start_column=1, start_row = 1+i, 
                end_column=end_col, end_row=1+i)

        self.sheet.cell(column=1, row=1).value = self.date_info.str_date
        self.sheet.cell(column=1, row=1).alignment = Alignment(horizontal="center", vertical="center")
        self.sheet.cell(column=1, row=2).value = str(self.selected_day)
        self.sheet.cell(column=1, row=2).alignment = Alignment(horizontal="center", vertical="center")

        y = self.data_start_y
        self.sheet.cell(column=1, row = y).value = "Подразделение"
        self.sheet.cell(column=1, row=y+1).value = "Кол-во человек"
        self.sheet.cell(column=1, row=y+2).value = "Персонал"
        
        self.sheet.row_dimensions[1].height = 30
        self.sheet.row_dimensions[2].height = 30
        self.sheet.row_dimensions[3].height = 40

        for x in range(1 + len(group_people_map) + 2):
            self.sheet.column_dimensions[get_column_letter(x+1)].width = 20

        for i, group_name in enumerate(sorted(list(group_people_map.keys()))):
            gx = self.data_start_x + i
            gy = self.data_start_y
            gcell = self.sheet.cell(column=gx, row=gy)
            gcell.value = group_name
            gcell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            people:list[PersonData] = group_people_map[group_name]
            ccell = self.sheet.cell(column=gx, row=gy+1)
            ccell.alignment = Alignment(horizontal='center', vertical='center')
            group_col = get_column_letter(gx)
            ccell.value = f'=COUNTIF({group_col}{gy+2}:{group_col}99,"*")'

            for p, person in enumerate(people):
                pcell = self.sheet.cell(column=gx, row=gy+2+p)
                pcell.alignment = Alignment(horizontal='left', vertical='center')
                pcell.value = person.name
        
        total_text_x = 2 + len(group_people_map)
        total_text_y = self.data_start_y
        tlabelcell = self.sheet.cell(column=total_text_x, row=total_text_y)
        tlabelcell.value = "всего человек:"
        tlabelcell.alignment = Alignment(horizontal='right', vertical='center')

        total_col_s = get_column_letter(2)
        total_col_e = get_column_letter(2 + len(group_people_map)-1)
        tcell = self.sheet.cell(column=total_text_x+1, row=total_text_y)
        tcell.value = f'=COUNTIF({total_col_s}{self.data_start_y+2}:{total_col_e}99,"*")'
        tcell.alignment = Alignment(horizontal='left', vertical='center')
    
    def make_report_sheet(self):
        self.data_start_x = 1
        self.data_start_y = 4
        group_people_map = self.group_people_map

        end_col = 3
        for i in range(2):
            self.sheet.merge_cells(
                start_column=1, start_row = 1+i, 
                end_column=end_col, end_row=1+i)

        self.sheet.cell(column=1, row=1).value = self.date_info.str_date
        self.sheet.cell(column=1, row=1).alignment = Alignment(horizontal="center", vertical="center")
        self.sheet.cell(column=1, row=2).value = str(self.selected_day)
        self.sheet.cell(column=1, row=2).alignment = Alignment(horizontal="center", vertical="center")

        x = self.data_start_x
        y = self.data_start_y-1
        self.sheet.cell(column=x,   row=y).value = "Подразделение"
        self.sheet.cell(column=x+1, row=y).value = "Кол-во"
        self.sheet.cell(column=x+2, row=y).value = "Персонал"
        
        self.sheet.row_dimensions[1].height = 20
        self.sheet.row_dimensions[2].height = 20

        widths = [32, 11, 100, 15, 7]
        for i, w in enumerate(widths):
            x = self.data_start_x + i
            self.sheet.column_dimensions[get_column_letter(x)].width = w

        for y in range(len(group_people_map)):
            y = self.data_start_y + y
            self.sheet.row_dimensions[y].height = 90

        total_people_count = 0

        for i, group_name in enumerate(sorted(list(group_people_map.keys()))):
            gx = self.data_start_x
            gy = self.data_start_y + i
            gcell = self.sheet.cell(column=gx, row=gy)
            gcell.value = group_name
            gcell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            people:list[PersonData] = group_people_map[group_name]
            ccell = self.sheet.cell(column=gx+1, row=gy)
            ccell.alignment = Alignment(horizontal='center', vertical='center')
            ccell.value = len(group_people_map[group_name])
            total_people_count += len(group_people_map[group_name])

            pcell = self.sheet.cell(column=gx+2, row=gy)
            pcell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            pcell.value = ", ".join(map(lambda x: x.name, people))
        
        total_text_x = self.data_start_x+3
        total_text_y = self.data_start_y-1
        tlabelcell = self.sheet.cell(column=total_text_x, row=total_text_y)
        tlabelcell.value = "всего человек:"
        tlabelcell.alignment = Alignment(horizontal='right', vertical='center')

        tcell = self.sheet.cell(column=total_text_x+1, row=total_text_y)
        tcell.value = total_people_count
        tcell.alignment = Alignment(horizontal='left', vertical='center')
        

        