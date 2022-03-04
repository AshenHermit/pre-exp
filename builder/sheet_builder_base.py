from pathlib import Path
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl import load_workbook
from openpyxl.styles.alignment import Alignment
from openpyxl.styles.fonts import Font
from openpyxl.utils import get_column_letter
import typing

from .data import RulesLibrary, StaffData
from .data_readers import DateInfo, DateReader, RulesReader, StaffDataReader
import datetime

class ExcelSheetBuilder():
    def __init__(self, data_sheet:Worksheet=None, sheet:Worksheet=None):
        self.data_sheet = data_sheet
        self.sheet = sheet
        self.workbook:Workbook = None
        
        self.staff:StaffData = None
        self.rules:RulesLibrary = None
        self.date_info:DateInfo = None

        self.data_start_x = 5
        self.data_start_y = 4

        self.days_count_in_month = 31
        self.last_cell_x = self.data_start_x + self.days_count_in_month

    def setup_data(self):
        self.staff:StaffData = StaffDataReader(self.data_sheet).read_staff_data()
        self.rules:RulesLibrary = RulesReader(self.data_sheet).read_rule_library()
        self.date_info = DateReader(self.data_sheet).read_date_info()

    def build(self):
        self.setup_data()
        self.render_base()
        self.fill_data()

    def render_base(self):
        pass

    def render_people_list(self, name_repeat=1):
        self.last_cell_x = 1 + self.days_count_in_month+(2+1+1+1)

        self.sheet.merge_cells('A1:A2')
        self.sheet.merge_cells('B1:C2')
        self.sheet.merge_cells('D1:D2')
        self.sheet.merge_cells(start_row=1, start_column=5, end_row=1, end_column=self.last_cell_x)

        # staff
        for i, emp in enumerate(self.staff.people):

            for rp in range(name_repeat):
                y = i*2 * name_repeat + rp*2

                y_offset = self.data_start_y
                
                for x in range(1, 4):
                    self.sheet.merge_cells(start_row=y_offset + y, start_column=x, end_row=y_offset + y+1, end_column=x)
            
                # name
                self.sheet.cell(column=1, row=y_offset + y).alignment = Alignment(horizontal='left', vertical='center')
                self.sheet.cell(column=1, row=y_offset + y).value = emp.name

                # number
                self.sheet.cell(column=2, row=y_offset + y).alignment = Alignment(horizontal='center', vertical='center')
                self.sheet.cell(column=2, row=y_offset + y).value = i+1

                # job
                self.sheet.cell(column=4, row=y_offset + y).alignment = Alignment(horizontal='center', vertical='center')
                self.sheet.cell(column=4, row=y_offset + y).value = emp.job
                self.sheet.cell(column=4, row=y_offset + y+1).alignment = Alignment(horizontal='center', vertical='center')
                self.sheet.cell(column=4, row=y_offset + y+1).value = 1


        for x in range(self.data_start_x, self.last_cell_x):
            self.sheet.column_dimensions[get_column_letter(x)].width = 6.4

    def fill_data(self):
        pass

    def format_cell_value(self, value):
        if type(value) is datetime.time:
            value = value.strftime("%H:%M")
        elif type(value) is float:
            value = round(value*100)/100
        elif type(value) is int:
            pass
        elif type(value) is not str:
            value = str(value)
        return value