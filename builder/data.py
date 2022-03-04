from functools import cache
import shlex
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl import load_workbook
from openpyxl.styles.alignment import Alignment
from openpyxl.styles.fonts import Font
from openpyxl.utils import get_column_letter
import typing

# staff
class PersonData():
    def __init__(self, name:str, job:str, days:list):
        self.name:str = name
        self.job:str = job
        self.days:list[str] = days

class StaffData():
    def __init__(self):
        self.people:typing.List[PersonData] = []
    
    def sort(self):
        self.people = sorted(self.people, key=lambda x: x.name)

    def add_person(self, person:PersonData):
        self.people.append(person)


# rules
class SimpleTable():
    def __init__(self) -> None:
        self.values = {}
        self.size = [0, 0]

    @staticmethod
    def _pos_to_key(x:int, y:int):
        return f"{x},{y}"

    def get(self, x:int, y:int, default=""):
        key = self._pos_to_key(x, y)
        return self.values.get(key, default)

    def set(self, x:int, y:int, value=""):
        if value is None: return
        key = self._pos_to_key(x, y)
        self.values[key] = value
        self.size[0] = max(self.size[0], x+1)
        self.size[1] = max(self.size[1], y+1)

    @staticmethod
    def from_sheet_region(sheet:Worksheet, region_start_pos:typing.Tuple[int, int], region_end_pos:typing.Tuple[int, int]):
        table = SimpleTable()
        start, end = region_start_pos, region_end_pos
        for i in range(end[0] - start[0]):
            for j in range(end[1] - start[1]):
                value = sheet.cell(column=start[0]+i, row=start[1]+j).value
                table.set(i, j, value)
        return table

    def iterate(self) -> typing.Generator[int, int, object]:
        for x in range(self.size[0]):
            for y in range(self.size[1]):
                yield x, y, self.get(x, y, "")

class Rule():
    def __init__(self) -> None:
        self.day_values:list[str] = []
        self.day_values:list[str] = []
        self.tables:list[SimpleTable] = []

    def test_day(self, day: str):
        day = str(day).lower().strip()

        for pat in self.day_values:
            pat = pat.strip().lower()
            if day == pat:
                return True
        return False

    def test_job(self, job: str):
        job = str(job).lower().strip()

        table = self.tables[2]
        job_patterns = self.job_patterns
        if len(job_patterns)==0: return True
        
        for pat in job_patterns:
            pat = pat.strip().lower()
            if job.find(pat) != -1: return True
        return False

    @property
    def job_patterns(self):
        table = self.tables[2]
        job_patterns = []
        for i in range(table.size[1]):
            patterns = shlex.split(str(table.get(0, i, "").replace(",", " ")).lower())
            job_patterns += patterns
        return job_patterns
    
    @property
    def groups_names(self):
        groups_names = set()
        for i,j,value in self.tables[3].iterate():
            if value:
                groups_names.add(value)
        return list(groups_names)

class RulesLibrary():
    def __init__(self) -> None:
        self.rules:list[Rule] = []

    def get_fit_rule(self, day:str=None, job:str=None):
        day = day or ""
        job = job or ""
        fit_rule = None
        fit_value = 0
        for rule in self.rules:
            if rule.test_day(day) and rule.test_job(job):
                value = 0
                if len(rule.job_patterns)>0:
                    value+=1
                if value>=fit_value:
                    fit_value = value
                    fit_rule = rule
        return fit_rule