from pathlib import Path
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl import load_workbook
from openpyxl.styles.alignment import Alignment
from openpyxl.styles.fonts import Font
from openpyxl.utils import get_column_letter
import typing
import datetime

from builder.builder import ExcelSheetBuilder

from .data import RulesLibrary, StaffData
from .data_readers import RulesReader, StaffDataReader

class PrepaidExpenseBuilder(ExcelSheetBuilder):
    def render_base(self):
        self.render_people_list(name_repeat=2)
        self.render_titles()

    def render_titles(self):
        # main data
        for x in range(1, self.last_cell_x+1): # + Уч. номер(2) + Должность(1) + Итого_1(1) + Всего(1)
            self.sheet.cell(column=x, row=3).font = Font(size=6)

            self.sheet.cell(column=x, row=3).alignment = Alignment(horizontal='center')
            self.sheet.cell(column=x, row=3).value = x

            if x == 1:
                self.sheet.cell(column=x, row=1).alignment = Alignment(horizontal='center')
                self.sheet.cell(column=x, row=1).value = "Фамилия имя отчетсво\n(строго в алфавитном порядке)"

            elif x==2:  
                self.sheet.cell(column=x, row=1).alignment = Alignment(horizontal='center')
                self.sheet.cell(column=x, row=1).value = "Уч. номер"

            elif x==4:
                self.sheet.cell(column=x, row=1).alignment = Alignment(horizontal='center')
                self.sheet.cell(column=x, row=1).value = "Должность (профессия) / Ставка"

            elif x==5:
                self.sheet.cell(column=x, row=1).alignment = Alignment(horizontal='center')
                self.sheet.cell(column=x, row=1).value = "Числа месяца"

            if x>4 and x<20:
                self.sheet.cell(column=x, row=2).alignment = Alignment(horizontal='center')
                self.sheet.cell(column=x, row=2).value = x-4

            elif x==20:
                self.sheet.cell(column=x, row=2).alignment = Alignment(horizontal='center')
                self.sheet.cell(column=x, row=2).value = "Итого"

            elif x>20 and x < self.last_cell_x:
                self.sheet.cell(column=x, row=2).alignment = Alignment(horizontal='center')
                self.sheet.cell(column=x, row=2).value = x-5

            elif x==self.last_cell_x:
                self.sheet.cell(column=x, row=2).alignment = Alignment(horizontal='center')
                self.sheet.cell(column=x, row=2).value = "Всего"
    
    def fill_data(self):
        #staff
        for i, emp in enumerate(self.staff.people):
            # day / night
            for dn in range(2):
                y = i*4+dn*2

                # formulas
                self.sheet.cell(column=20, row=self.data_start_y + y).value = "=SUM(E{0}:S{0})".format(self.data_start_y + y)
                if self.days_count_in_month == 30:
                    self.sheet.cell(column=self.last_cell_x, row=self.data_start_y + y).value = "=SUM(U{0}:AI{0})+T{0}".format(self.data_start_y + y)
                if self.days_count_in_month == 31:
                    self.sheet.cell(column=self.last_cell_x, row=self.data_start_y + y).value = "=SUM(U{0}:AJ{0})+T{0}".format(self.data_start_y + y)

            y = i*4
            # days
            for x, day in reversed(list(enumerate(emp.days))):
                dx = self.data_start_x + x
                dy = self.data_start_y + y
                
                if day == None:
                    day = ""
                else:
                    self.set_table_value(dx, dy+1, day)

                rule = self.rules.get_fit_rule(day, emp.job)
                if rule is not None:
                    exp_table = rule.tables[0]
                    for i in range(exp_table.size[0]):
                        for j in range(exp_table.size[1]):
                            self.set_table_value(dx+i, dy+j, exp_table.get(i, j))

    def set_table_value(self, x, y, value, use_limit=True):
        value = self.format_cell_value(value)

        if x >= 20:
            x+=1
            limit = 1 + self.days_count_in_month + (2+1*3)
        else:
            limit = self.days_count_in_month + 5

        if x < limit or not use_limit:
            self.sheet.cell(column=x, row=y).alignment = Alignment(horizontal='center', vertical='center')
            self.sheet.cell(column=x, row=y).value = value